"""赛博瑞 RAG System CLI — data-driven content factory.

Commands:
  generate              Generate script from product brief (RAG-enhanced LLM)
  generate-storyboard   Finalized .docx script → storyboard .xlsx
  storyboard            Generate storyboard directly from product brief (RAG-enhanced)
  audit                 Audit script text or storyboard JSON for quality
  search                Semantic search the knowledge base
  stats                 Knowledge base statistics
  competitive search    Search Bilibili and run full competitive video analysis
  competitive report    Generate formatted .docx competitive analysis report
"""

import click
from rag_system.utils import logger

# ---- Input validation ----

VALID_CATEGORIES = {
    "keyboard", "monitor", "mouse", "gpu", "laptop", "headphone",
    "phone", "desk_chair", "speaker",
}
VALID_PERSONAS = {"折腾到吐", "好设牛啊", "朋克", "超机懂"}
VALID_FORMATS = {"review", "tierlist", "comparison", "hkrr", "hamd"}
VALID_MODES = {"normal", "experimental"}


def _validate_category(category: str):
    if category not in VALID_CATEGORIES:
        close = min(VALID_CATEGORIES, key=lambda c: _edit_distance(c, category))
        msg = f"未知品类 '{category}'。你是想说 '{close}' 吗？\n可选: {', '.join(sorted(VALID_CATEGORIES))}"
        raise click.BadParameter(msg, param_hint="--category / -c")


def _validate_persona(persona: str):
    if persona not in VALID_PERSONAS:
        raise click.BadParameter(
            f"未知人设 '{persona}'。可选: {', '.join(sorted(VALID_PERSONAS))}",
            param_hint="--persona")


def _validate_format(fmt: str):
    if fmt not in VALID_FORMATS:
        raise click.BadParameter(
            f"未知格式 '{fmt}'。可选: {', '.join(sorted(VALID_FORMATS))}",
            param_hint="--format")


def _validate_mode(mode: str):
    if mode not in VALID_MODES:
        raise click.BadParameter(
            f"未知模式 '{mode}'。可选: {', '.join(sorted(VALID_MODES))}",
            param_hint="--mode")


def _edit_distance(a: str, b: str) -> int:
    """Levenshtein distance for typo suggestions."""
    if len(a) < len(b):
        a, b = b, a
    if len(b) == 0:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        curr = [i + 1]
        for j, cb in enumerate(b):
            curr.append(min(
                prev[j + 1] + 1, curr[j] + 1,
                prev[j] + (0 if ca == cb else 1)
            ))
        prev = curr
    return prev[-1]


@click.group()
def cli():
    """赛博瑞 RAG System — 数据驱动内容工厂.

    AI-powered video script generation, storyboarding, and competitive
    analysis pipeline for Douyin (TikTok) tech review content.
    """
    from rag_system.utils import setup_logging
    setup_logging()


# ============================================================
# generate — Script from product brief
# ============================================================

@cli.command("generate")
@click.option("--product", "-p", required=True, help="产品名称，如：ROG龙鳞ACE MINI")
@click.option("--category", "-c", required=True,
              help="品类：keyboard / mouse / monitor / laptop / phone / gpu / headphone / desk_chair")
@click.option("--key-points", "-k", default="", help="核心卖点，逗号分隔，如：轻量化54g, 8K回报率, 399元")
@click.option("--brief", "-b", default=None, type=click.Path(exists=True),
              help="Brief文档路径 (.txt)，自动解析卖点和封面建议")
@click.option("--persona", default="折腾到吐", help="人设名称 (默认: 折腾到吐)")
@click.option("--price", default="", help="价格信息")
@click.option("--competitors", default="", help="竞品信息")
@click.option("--duration", "-d", default=2.0, type=float, help="目标时长，分钟 (默认: 2.0)")
@click.option("--format", "script_format", default="review",
              help="脚本格式：review / tierlist / comparison / hkrr / hamd (默认: review)")
@click.option("--mode", default="normal",
              help="生成模式：normal / experimental（更大token、更激进风格）")
@click.option("--perspective", default="",
              help="个人观点注入 — 你对产品的真实体验，如：这个传感器暗光下提升明显")
@click.option("--temperature", default=0.8, type=float, help="LLM 温度 (默认: 0.8)")
@click.option("--output", "-o", default=None, type=click.Path(dir_okay=False), help="输出文件路径 (.txt)")
def generate(product, category, key_points, brief, persona, price, competitors,
             duration, script_format, mode, perspective, temperature, output):
    """Generate video script from product brief using RAG-enhanced LLM.

    Retrieves relevant past scripts from the knowledge base as style
    reference, then calls DeepSeek API to produce a ~800-1200 character
    Douyin-style review script.

    Examples:

        python -m rag_system generate -p "ROG龙鳞ACE MINI" -c mouse \\
            -k "54g轻量化,8K回报率,399元" --format tierlist

        python -m rag_system generate -p "迈从K20" -c speaker \\
            --brief briefs/maicong.txt
    """
    from pathlib import Path
    from rag_system.generation.generator import Generator
    from rag_system.retrieval.retriever import Retriever
    from rag_system.embedding.embedder import Embedder
    from rag_system.storage.vector_store import VectorStore

    # Validate: either --key-points or --brief must be provided
    if not key_points and not brief:
        raise click.UsageError("必须提供 --key-points 或 --brief")

    _validate_category(category)
    _validate_persona(persona)
    _validate_format(script_format)
    _validate_mode(mode)

    if perspective:
        from rag_system.generation.prompts import PERSPECTIVE_INJECTION

    click.echo(f"Generating script for: {product} [{category}]")

    # --- Brief Analyzer integration (封面前置) ---
    brief_context = ""
    cover_suggestion = ""

    if brief:
        from rag_system.generation.brief_analyzer import (
            parse_brief, brief_to_prompt_context, generate_recommendation,
        )
        from rag_system.utils import sanitize_filename

        brief_text = Path(brief).read_text(encoding="utf-8")
        analysis = parse_brief(brief_text)
        brief_context = brief_to_prompt_context(analysis)
        cover_suggestion = analysis.cover_suggestion

        # Derive key_points from brief if not explicitly provided
        if analysis.selling_points and not key_points:
            sp_names = [sp.name for sp in sorted(analysis.selling_points, key=lambda s: -s.priority)]
            key_points = ", ".join(sp_names[:6])
            click.echo(f"Parsed {len(analysis.selling_points)} selling points from brief")

        if analysis.must_mentions:
            click.echo(f"Must-mentions: {len(analysis.must_mentions)}")
        if cover_suggestion:
            click.echo(f"Cover concept: {cover_suggestion[:80]}...")

        # Save cover suggestion for cover command
        covers_dir = Path("output/covers")
        covers_dir.mkdir(parents=True, exist_ok=True)
        safe_name = sanitize_filename(product)
        cover_path = covers_dir / f"{safe_name}-cover-suggestion.txt"
        rec = generate_recommendation(analysis, persona=persona)
        cover_path.write_text(
            f"产品: {product}\n人设: {persona}\n品类: {category}\n\n"
            f"封面建议: {cover_suggestion}\n\n--- 完整编辑指南 ---\n{rec}",
            encoding="utf-8",
        )
        click.echo(f"Brief analysis saved: {cover_path}")

    # --- 封面前置: 封面方向注入 → 文案围绕封面视角展开 ---
    cover_direction = ""
    if cover_suggestion:
        cover_direction = (
            f"本次视频封面方向：{cover_suggestion}。"
            f"重要：请确保脚本开场hook能直接支撑这个封面标题，"
            f"全文内容围绕封面承诺展开，不要跑题。"
        )
        click.echo(f"Cover-first: 封面方向已注入prompt")

    # --- Analytics反哺: 查历史表现 + 审计失败项 → 指导本次生成 ---
    analytics_context = ""
    try:
        from rag_system.generation.analytics import read_events
        from collections import Counter
        events = read_events(days=90)

        # 1. 基础统计
        relevant = [e for e in events
                    if e.get("type") == "generate"
                    and e.get("persona") == persona
                    and e.get("category") == category]
        parts = []
        if relevant:
            avg_chars = sum(e.get("char_count", 0) for e in relevant) // len(relevant)
            formats_used = list(set(e.get("format", "review") for e in relevant))
            parts.append(
                f"历史：'{persona}'在'{category}'已生成{len(relevant)}次，"
                f"平均{avg_chars}字/篇，常用{', '.join(formats_used[:3])}格式"
            )

        # 2. 审计失败项挖掘 — 这才是真正的"学习"
        audit_events = [e for e in events
                        if e.get("type") == "audit"
                        and e.get("persona") == persona
                        and e.get("category") == category]
        if audit_events:
            failed_counter = Counter()
            for ae in audit_events:
                for fname in ae.get("failed_checks", []):
                    failed_counter[fname] += 1

            if failed_counter:
                # Map audit check names to actionable writing tips
                FIX_TIPS = {
                    "口语化程度": "多用短句和语气词（吧、啊、呢），每句不超过25字，像在跟朋友聊天",
                    "电商味": "避免'限时抢购''手慢无''全网最低'等电商促使用语，用体验描述代替",
                    "态度密度": "每段至少1处明确态度——'有一说一''说实话''我个人觉得'",
                    "禁用词": "检查禁用词清单，替换为口语化表达",
                    "流水账检测": "避免'首先/然后/接着/最后'的流水账结构，每段用钩子开场",
                    "长短句节奏": "长短句交替——长句讲道理(≤25字)，短句给结论(≤10字)，比例约2:1",
                    "价格检测": "价格信息自然融入体验描述中，不要单独报价",
                    "卖点覆盖": "确保每个核心卖点都有对应口播段落，不遗漏",
                }
                top_fails = failed_counter.most_common(3)
                tips = []
                for fname, count in top_fails:
                    tip = FIX_TIPS.get(fname, f"注意'{fname}'问题")
                    tips.append(f"· {tip}（历史失败{count}次）")
                if tips:
                    parts.append("历史高频失败项，务必注意：\n" + "\n".join(tips))

        if parts:
            analytics_context = "\n".join(parts)
            click.echo(f"Analytics: {len(relevant)} refs, {len(audit_events)} audit records")
    except Exception as e:
        logger.warning("Non-critical error: %s", e)

    # RAG retrieval for style context
    embedder = Embedder()
    store = VectorStore()
    retriever = Retriever(embedder, store)
    query = f"{product} {category} {key_points}"
    chunks = retriever.retrieve(query, top_k=8, category=category)

    if chunks:
        click.echo(f"Retrieved {len(chunks)} reference chunks from knowledge base")

    gen = Generator()
    script = gen.generate(
        product_name=product,
        category=category,
        key_points=key_points,
        persona=persona,
        price=price,
        competitors=competitors,
        duration_minutes=duration,
        script_format=script_format,
        retrieved_chunks=chunks,
        temperature=temperature,
        brief_context=brief_context,
        cover_direction=cover_direction,
        analytics_context=analytics_context,
        mode=mode,
        perspective_context=(
            PERSPECTIVE_INJECTION.format(perspectives=perspective)
            if perspective else ""
        ),
    )

    if output:
        path = Path(output)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(script, encoding="utf-8")
        click.echo(f"Script saved to: {path}")
    else:
        click.echo("\n" + "=" * 60)
        click.echo(script)
        click.echo("=" * 60)

    click.echo(f"Done: {len(script)} characters")

    # Output registration
    if output:
        from rag_system.generation.output_manager import register_output
        register_output("script", Path(output), {
            "product": product, "persona": persona, "category": category,
            "format": script_format,
        })

    # Pipeline analytics event
    try:
        from rag_system.generation.analytics import log_event
        log_event("generate", product=product, persona=persona, category=category,
                  char_count=len(script), format=script_format, wiki_used=bool(chunks))
    except Exception as e:
        logger.warning("Non-critical error: %s", e)

    # Audit + log results for learning feedback loop
    try:
        from rag_system.generation.auditor import audit_script
        audit_result = audit_script(script, key_points=key_points, duration_minutes=duration)
        failed = [c["name"] for c in audit_result.checks if not c.get("passed")]
        passed_n = sum(1 for c in audit_result.checks if c.get("passed"))
        total_n = len(audit_result.checks)
        click.echo(f"Audit: {passed_n}/{total_n} passed" + (f" | Failed: {', '.join(failed)}" if failed else ""))
        log_event("audit", product=product, persona=persona, category=category,
                  passed=audit_result.passed, total_checks=total_n,
                  passed_count=passed_n, failed_checks=failed,
                  warnings=len(audit_result.warnings))
    except Exception as e:
        logger.warning("Non-critical error: %s", e)


# ============================================================
# generate-storyboard — Finalized script → storyboard xlsx
# ============================================================

@cli.command("generate-storyboard")
@click.argument("script", type=click.Path(exists=True))
@click.argument("product")
@click.argument("persona", default="折腾到吐")
@click.option("--format-ref", default=None, type=click.Path(exists=True),
              help="甲方参考xlsx文件，自动匹配列格式输出")
@click.option("--columns", default=None,
              help="逗号分隔的列名，如：镜头,时间,画面描述,口播,备注")
@click.option("--preview", is_flag=True, default=False,
              help="仅预览列映射和样例行，不调LLM生成")
def generate_storyboard(script: str, product: str, persona: str,
                        format_ref: str | None = None, columns: str | None = None,
                        preview: bool = False):
    """Convert finalized .docx script into a shot-by-shot storyboard .xlsx.

    Pipeline: parse .docx → LLM shot breakdown → audit → auto-fix → save.

    SCRIPT: Path to the finalized .docx script file.
    PRODUCT: Product name for the storyboard title.
    PERSONA: Persona name (default: 折腾到吐).

    Examples:

        python -m rag_system generate-storyboard output/scripts/ROG.docx "ROG龙鳞ACE MINI"

        python -m rag_system generate-storyboard output/scripts/ROG.docx "ROG" \\
            --columns "镜头,时间,画面描述,口播,备注"

        python -m rag_system generate-storyboard output/scripts/ROG.docx "ROG" \\
            --columns "镜头,时间,画面描述,口播,备注" --preview

        python -m rag_system generate-storyboard output/scripts/ROG.docx "ROG" \\
            --format-ref 甲方参考.xlsx --preview
    """
    from pathlib import Path
    from rag_system.generation.script_to_storyboard import storyboard_pipeline, parse_docx_script

    ref_path = Path(format_ref) if format_ref else None
    col_list = [c.strip() for c in columns.split(",") if c.strip()] if columns else None

    # --- Preview mode: show mapping + sample, skip LLM ---
    if preview:
        from rag_system.generation.template_adapter import (
            preview_column_mapping, build_column_mapping,
        )
        from openpyxl import load_workbook

        # Resolve columns from reference file or explicit list or default
        if ref_path:
            if ref_path.suffix in ('.xlsx', '.xlsm'):
                wb = load_workbook(str(ref_path))
                ws = wb.active
                from rag_system.generation.template_adapter import detect_header_row
                hr = detect_header_row(ws)
                col_list = [str(ws.cell(row=hr, column=c).value or "").strip().replace("\n", " ")
                           for c in range(1, (ws.max_column or 20) + 1)]
                col_list = [c for c in col_list if c]
                wb.close()
                click.echo(f"从参考文件读取 {len(col_list)} 列")
            else:
                click.echo("参考文件不是xlsx格式，无法预览列映射", err=True)
                return
        elif not col_list:
            col_list = ["镜号", "景别·运镜", "画面描述", "口播文案", "时长",
                       "花字/特效", "音效/声画", "灯光/机位", "备注"]
            click.echo("使用默认9列格式")

        # Parse script for sample VO
        script_data = parse_docx_script(Path(script))
        sample_vo = (script_data.get("body", [""]) or [""])[0][:80] if script_data.get("body") else ""
        sample_shot = {
            "shot_number": 1, "duration": "3", "visual": "产品主图+外观特写",
            "voiceover": sample_vo, "jingbie": "中景", "yunjing": "推",
            "huazi": "开场钩子", "audio": "", "lighting": "", "camera_setup": "", "notes": ""
        }

        click.echo(preview_column_mapping(col_list, sample_shot))
        click.echo("[OK] 确认格式无误？去掉 --preview 即可全量生成。")
        return

    click.echo(f"Generating storyboard for: {product}")
    result = storyboard_pipeline(Path(script), product, persona,
                                 reference_path=ref_path, columns=col_list)
    click.echo(f"Done: {result}")
    from rag_system.generation.output_manager import register_output
    register_output("storyboard", result, {"product": product, "persona": persona})


# ============================================================
# storyboard — RAG storyboard from product brief (no script step)
# ============================================================

@cli.command("storyboard")
@click.option("--product", "-p", required=True, help="产品名称")
@click.option("--category", "-c", required=True,
              help="品类：keyboard / mouse / monitor / laptop / phone / gpu / headphone / desk_chair")
@click.option("--key-points", "-k", required=True, help="核心卖点，逗号分隔")
@click.option("--persona", default="折腾到吐", help="人设名称 (默认: 折腾到吐)")
@click.option("--price", default="", help="价格信息")
@click.option("--competitors", default="", help="竞品信息")
@click.option("--extra-notes", default="", help="额外备注（拍摄要求、风格偏好等）")
@click.option("--temperature", default=0.8, type=float, help="LLM 温度 (默认: 0.8)")
@click.option("--output", "-o", default=None, type=click.Path(file_okay=False),
              help="输出目录 (默认: output/storyboards)")
@click.option("--no-audit", is_flag=True, default=False, help="跳过自审步骤")
@click.option("--format-ref", default=None, type=click.Path(exists=True),
              help="甲方参考xlsx文件，自动匹配列格式输出")
@click.option("--columns", default=None,
              help="逗号分隔的列名，如：镜头,时间,画面描述,口播,备注")
@click.option("--preview", is_flag=True, default=False,
              help="仅预览列映射，不调LLM生成")
def storyboard(product, category, key_points, persona, price, competitors,
               extra_notes, temperature, output, no_audit, format_ref, columns,
               preview=False):
    """Generate storyboard directly from product brief (RAG-enhanced).

    Skips the script step — goes straight from brief to camera-ready
    shot list. Retrieves past storyboards from the knowledge base as
    stylistic reference.

    Examples:

        python -m rag_system storyboard -p "ROG龙鳞" -c mouse \\
            -k "54g,PAW3950,399元"

        python -m rag_system storyboard -p "迈从K20" -c speaker \\
            -k "双声道,DSP芯片,99元" --no-audit
    """
    from collections import Counter
    from pathlib import Path

    from rag_system.generation.storyboard_generator import StoryboardGenerator, ProductBrief
    from rag_system.generation.xlsx_formatter import format_storyboard_to_xlsx
    from rag_system.retrieval.retriever import Retriever
    from rag_system.embedding.embedder import Embedder
    from rag_system.storage.vector_store import VectorStore
    from rag_system.utils import sanitize_filename

    # --- Preview mode ---
    if preview:
        from rag_system.generation.template_adapter import preview_column_mapping
        from openpyxl import load_workbook

        ref_path = Path(format_ref) if format_ref else None
        col_list = [c.strip() for c in columns.split(",") if c.strip()] if columns else None

        if ref_path:
            if ref_path.suffix in ('.xlsx', '.xlsm'):
                wb = load_workbook(str(ref_path))
                ws = wb.active
                from rag_system.generation.template_adapter import detect_header_row
                hr = detect_header_row(ws)
                col_list = [str(ws.cell(row=hr, column=c).value or "").strip().replace("\n", " ")
                           for c in range(1, (ws.max_column or 20) + 1)]
                col_list = [c for c in col_list if c]
                wb.close()
                click.echo(f"从参考文件读取 {len(col_list)} 列")
            else:
                click.echo("参考文件不是xlsx格式，无法预览列映射", err=True)
                return
        elif not col_list:
            col_list = ["镜号", "景别·运镜", "画面描述", "口播文案", "时长",
                       "花字/特效", "音效/声画", "灯光/机位", "备注"]
            click.echo("使用默认9列格式")

        sample_shot = {
            "shot_number": 1, "duration": "3", "visual": "产品主图+外观特写",
            "voiceover": f"{product} {category} 评测。{key_points[:40]}...",
            "jingbie": "中景", "yunjing": "推",
            "huazi": key_points[:20], "audio": "", "lighting": "", "camera_setup": "", "notes": ""
        }
        click.echo(preview_column_mapping(col_list, sample_shot))
        click.echo("[OK] 确认格式无误？去掉 --preview 即可全量生成。")
        return

    click.echo(f"Generating storyboard from brief: {product} [{category}]")

    # RAG retrieval for style context
    embedder = Embedder()
    store = VectorStore()
    retriever = Retriever(embedder, store)
    query = f"{product} {category} {key_points}"
    chunks = retriever.retrieve(query, top_k=8, category=category)

    if chunks:
        click.echo(f"Retrieved {len(chunks)} reference chunks from knowledge base")

    brief = ProductBrief(
        product_name=product,
        category=category,
        persona=persona,
        key_points=key_points,
        price=price,
        competitors=competitors,
        extra_notes=extra_notes,
    )

    gen = StoryboardGenerator()
    result = gen.generate(brief, retrieved_chunks=chunks, temperature=temperature)
    shots = result.get("shots", [])

    # Optional audit
    if not no_audit and shots:
        from rag_system.generation.auditor import audit_storyboard, audit_shootability
        click.echo("Running self-audit...")
        audit_result = audit_storyboard(result, key_points=key_points)
        passed = sum(1 for c in audit_result.checks if c["passed"])
        total = len(audit_result.checks)
        click.echo(f"Content audit: {passed}/{total} checks passed")
        if audit_result.warnings:
            click.echo(f"Warnings: {len(audit_result.warnings)}")
        if audit_result.suggestions:
            click.echo(f"Suggestions: {len(audit_result.suggestions)}")

    # Save to xlsx
    output_dir = Path(output) if output else Path("output/storyboards")
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_name = sanitize_filename(product)
    import time as _time
    safe_path = output_dir / f"{safe_name}-{persona}-分镜表_{int(_time.time()) % 100000}.xlsx"
    col_list = [c.strip() for c in columns.split(",") if c.strip()] if columns else None
    format_storyboard_to_xlsx(
        storyboard=result,
        product_name=product,
        persona=persona,
        output_path=safe_path,
        reference_path=Path(format_ref) if format_ref else None,
        columns=col_list,
    )

    # Summary
    jingbies = Counter(s.get("jingbie", "") for s in shots)
    yunjings = Counter(s.get("yunjing", "") for s in shots)
    huazi_shots = sum(1 for s in shots if s.get("huazi", "").strip())
    trans_shots = sum(1 for s in shots if s.get("transition", "") not in ("硬切", "开场", ""))
    total_vo = sum(len(s.get("voiceover", "")) for s in shots)

    click.echo(f"\n{'='*50}")
    click.echo(f"  Storyboard: {product} — {persona}")
    click.echo(f"  Shots: {len(shots)} | VO: {total_vo} chars | Transitions: {trans_shots}")
    click.echo(f"  Huazi: {huazi_shots} shots | Jingbie: {dict(jingbies.most_common(4))}")
    click.echo(f"  File: {safe_path}")
    click.echo(f"{'='*50}")

    # Pipeline analytics event
    try:
        from rag_system.generation.analytics import log_event
        log_event("storyboard", product=product, persona=persona, category=category,
                  shot_count=len(shots), vo_chars=total_vo,
                  yunjing_variety=len(set(s.get("yunjing", "") for s in shots)))
    except Exception as e:
        logger.warning("Non-critical error: %s", e)


# ============================================================
# audit — Audit script text or storyboard JSON
# ============================================================

@cli.command("audit")
@click.argument("input_file", type=click.Path(exists=True))
@click.option("--key-points", "-k", default="", help="核心卖点(逗号分隔)，用于检查卖点覆盖")
@click.option("--duration", "-d", default=2.0, type=float, help="目标时长，分钟 (默认: 2.0)")
@click.option("--audience", is_flag=True, default=False,
              help="AI点映团 — 模拟3种观众审稿，出无趣时间表")
def audit(input_file, key_points, duration, audience):
    """Audit script or storyboard for quality issues.

    Auto-detects format:
      .json  → audits as storyboard (shot count, transitions, etc.)
      .docx  → parses body text and audits as script
      .txt   → audits as plain-text script

    Checks include: forbidden words, e-commerce smell, spoken language
    density, attitude density, sentence rhythm, selling-point coverage,
    and (for storyboards) shot variety, transitions, and shootability.

    Add --audience for AI preview audience review (影视飓风 AI点映团).

    Examples:

        python -m rag_system audit output/scripts/ROG.txt -k "54g,8K,399元"

        python -m rag_system audit output/storyboards/ROG-分镜表.json
    """
    import json
    from pathlib import Path

    input_path = Path(input_file)

    # For audience review, capture the script text
    audience_script = ""

    if input_path.suffix == ".json":
        # Audit as storyboard
        from rag_system.generation.auditor import audit_storyboard
        try:
            storyboard_data = json.loads(input_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            click.echo(f"Error: Invalid JSON in {input_file}", err=True)
            raise SystemExit(1)
        result = audit_storyboard(storyboard_data, key_points=key_points, duration_minutes=duration)
        click.echo(f"\nAuditing storyboard: {input_path.name}")
        # Extract VO for audience review
        audience_script = "\n".join(s.get("voiceover", "") for s in storyboard_data.get("shots", []))
    elif input_path.suffix == ".docx":
        # Parse docx body and audit as script
        from rag_system.generation.script_to_storyboard import parse_docx_script
        from rag_system.generation.auditor import audit_script
        script_data = parse_docx_script(input_path)
        audience_script = script_data["full_script"]
        result = audit_script(audience_script, key_points=key_points, duration_minutes=duration)
        click.echo(f"\nAuditing script (docx): {input_path.name}")
    else:
        # Audit as plain text script
        from rag_system.generation.auditor import audit_script
        audience_script = input_path.read_text(encoding="utf-8")
        result = audit_script(audience_script, key_points=key_points, duration_minutes=duration)
        click.echo(f"\nAuditing script: {input_path.name}")

    click.echo(result.summarize())

    if result.passed:
        click.echo("\n✓ Audit passed")
    else:
        click.echo("\n✗ Audit found issues — review warnings and suggestions above")

    # AI点映团 — 多视角审稿
    if audience:
        from rag_system.generation.preview_audience import (
            run_audience_review, format_audience_review,
        )
        product_name = input_path.stem.split("-")[0][:30]
        click.echo(f"\n{'='*50}")
        click.echo(f"  AI点映团审稿 — {product_name}")
        click.echo(f"{'='*50}")
        try:
            review = run_audience_review(
                script=audience_script,
                product_name=product_name,
                category="",
                persona="",
            )
            click.echo(format_audience_review(review))
        except Exception as e:
            click.echo(f"AI点映团审稿失败: {e}")


# ============================================================
# search — Semantic search the knowledge base
# ============================================================

@cli.command("search")
@click.argument("query")
@click.option("--top-k", "-n", default=8, type=int, help="返回结果数量 (默认: 8)")
@click.option("--persona", default=None, help="按人设过滤：折腾到吐 / 等等")
@click.option("--category", "-c", default=None,
              help="按品类过滤：keyboard / mouse / monitor / laptop / phone / gpu / headphone / desk_chair")
@click.option("--final-only/--no-final-only", default=False, help="只返回定稿版本")
@click.option("--exclude-revisions/--no-exclude-revisions", default=False, help="排除修订版本")
def search(query, top_k, persona, category, final_only, exclude_revisions):
    """Semantic search the knowledge base for relevant script chunks.

    Searches across all ingested scripts and competitive analyses.
    Returns chunks ranked by cosine similarity, with metadata about
    source file, persona, category, and product.

    Examples:

        python -m rag_system search "磁轴键盘 手感" -c keyboard -n 5

        python -m rag_system search "ROG 轻量化 鼠标" --persona "折腾到吐" --final-only
    """
    from rag_system.retrieval.retriever import Retriever
    from rag_system.embedding.embedder import Embedder
    from rag_system.storage.vector_store import VectorStore

    embedder = Embedder()
    store = VectorStore()
    retriever = Retriever(embedder, store)

    click.echo(f'Searching: "{query}"')
    results = retriever.retrieve(
        query=query,
        top_k=top_k,
        persona=persona,
        category=category,
        is_final_only=final_only,
        exclude_revisions=exclude_revisions,
    )

    if not results:
        click.echo("No results found.")
        return

    click.echo(f"\nFound {len(results)} results:\n")
    for i, r in enumerate(results, 1):
        click.echo(f"--- Result {i} (score: {r.score:.4f}) ---")
        click.echo(f"  Source:   {r.source_file}")
        click.echo(f"  Persona:  {r.persona}  |  Category: {r.category}  |  Product: {r.product_name}")
        preview = r.document[:200].replace("\n", " ")
        click.echo(f"  Preview:  {preview}...")
        click.echo()


# ============================================================
# stats — Knowledge base statistics
# ============================================================

@cli.command("stats")
def stats():
    """Display knowledge base statistics.

    Shows total chunks, unique sources, distribution by persona
    and category, final version count, and competitive entries.

    Example:

        python -m rag_system stats
    """
    from collections import Counter
    from rag_system.storage.vector_store import VectorStore

    store = VectorStore()
    total = store.count()

    click.echo(f"\n{'='*50}")
    click.echo(f"  Knowledge Base Statistics")
    click.echo(f"{'='*50}")
    click.echo(f"  Total chunks: {total}")

    if total == 0:
        click.echo("  (Empty — ingest scripts first, then try again)")
        click.echo(f"{'='*50}\n")
        return

    metadatas = store.get_all_metadata()
    if not metadatas:
        click.echo("  No metadata available")
        click.echo(f"{'='*50}\n")
        return

    personas = Counter(m.get("persona", "unknown") for m in metadatas)
    categories = Counter(m.get("category", "unknown") for m in metadatas)
    source_files = set(m.get("source_file", "") for m in metadatas)
    finals = sum(1 for m in metadatas if m.get("is_final") == "true")
    competitive = sum(1 for m in metadatas if m.get("is_competitive") == "true")

    click.echo(f"  Unique source files: {len(source_files)}")
    click.echo(f"  Final versions:      {finals}")
    click.echo(f"  Competitive entries: {competitive}")
    click.echo(f"\n  By Persona:")
    for persona_name, count in personas.most_common(10):
        bar = "█" * min(count, 40)
        click.echo(f"    {persona_name:<20} {count:>5}  {bar}")
    click.echo(f"\n  By Category:")
    for cat, count in categories.most_common(12):
        bar = "█" * min(count, 40)
        click.echo(f"    {cat:<20} {count:>5}  {bar}")
    click.echo(f"{'='*50}\n")


# ============================================================
# init — First-run setup wizard
# ============================================================

@cli.command("init")
def init():
    """First-run setup wizard — check environment, configure API, initialize.

    Guides new users through Python version check, dependency verification,
    API key setup, directory creation, and vector store initialization.

    Example:

        python -m rag_system init
    """
    from rag_system.init_wizard import run_init
    success = run_init()
    if not success:
        raise SystemExit(1)


# ============================================================
# competitive — Competitive analysis commands
# ============================================================

@cli.group("competitive")
def competitive():
    """Competitive video analysis — learn from top creators.

    Search Bilibili for top-performing videos by category, download,
    transcribe, and analyze their script structure, hooks, and visual
    patterns. Results are stored for use in script generation and
    compiled into professional reports.
    """
    pass


@competitive.command("search")
@click.option("--category", "-c", required=True,
              help="品类：keyboard / mouse / monitor / laptop / phone / gpu / headphone / desk_chair")
@click.option("--top-n", "-n", default=3, type=int, help="分析的视频数量 (默认: 3)")
@click.option("--skip-download/--no-skip-download", default=False, help="跳过视频下载，使用缓存的音频/转录")
def competitive_search(category, top_n, skip_download):
    """Search Bilibili and run full competitive analysis pipeline.

    Downloads top videos by category, transcribes audio (Whisper),
    analyzes script hooks, spoken density, attitude patterns, and
    visual rhythm. Results are indexed into the knowledge base for
    RAG retrieval during script generation.

    Examples:

        python -m rag_system competitive search -c keyboard -n 5

        python -m rag_system competitive search -c mouse --skip-download
    """
    from rag_system.competitive.pipeline import run_pipeline

    click.echo(f"Starting competitive analysis: {category} (top {top_n})")
    if skip_download:
        click.echo("Download skipped — using cached data if available")

    results = run_pipeline(category=category, top_n=top_n, skip_download=skip_download)

    if not results:
        click.echo("No results. Check network connectivity or try a different category.")
        return

    click.echo(f"\nAnalysis complete: {len(results)} videos analyzed\n")
    for r in results:
        click.echo(f"  [{r.get('hook_type', 'N/A')}] {r['title'][:60]}")
        click.echo(f"       {r['creator']}  |  {r['views']:,} views")


@competitive.command("report")
@click.option("--period", default="weekly",
              help="报告周期：weekly / monthly (默认: weekly)")
@click.option("--category", "-c", default=None,
              help="只报告指定品类（默认：全部）")
@click.option("--output", "-o", default=None, type=click.Path(dir_okay=False),
              help="输出路径 (.docx)")
def competitive_report(period, category, output):
    """Generate a professionally formatted .docx competitive analysis report.

    Compiles all analyzed videos into a Swiss-design report with:
    creator rankings, hook-type trends, category insights, deep
    analysis callouts, and actionable recommendations.

    Examples:

        python -m rag_system competitive report

        python -m rag_system competitive report --period monthly -c keyboard
    """
    from rag_system.competitive.reporter import generate_docx_report
    from rag_system.competitive.store import get_analyzed_videos

    videos = get_analyzed_videos(category=category or "")

    if not videos:
        click.echo("No analyzed videos found. Run 'competitive search' first.")
        return

    click.echo(f"Generating {period} competitive report from {len(videos)} videos...")
    if category:
        click.echo(f"Filtered to category: {category}")

    path = generate_docx_report(videos=videos, period=period)
    click.echo(f"Report saved: {path}")


# ============================================================
# analytics — Pipeline analytics and reporting
# ============================================================

@cli.command("analytics")
@click.option("--days", "-d", default=30, type=int, help="统计周期，天 (默认: 30)")
@click.option("--persona", "-p", default=None, help="按人设过滤")
@click.option("--output", "-o", default=None, type=click.Path(dir_okay=False),
              help="输出JSON路径 (默认: 只打印)")
@click.option("--matrix", is_flag=True, default=False,
              help="人设×品类交叉效能矩阵 — 数据驱动人设选择")
def analytics(days, persona, output, matrix):
    """Show pipeline analytics report — 管线产量与效率分析.

    Tracks every 'generate' and 'storyboard' run with structured metrics.
    Shows persona breakdown, category distribution, format mix, and velocity.

    Examples:

        python -m rag_system analytics

        python -m rag_system analytics --days 7

        python -m rag_system analytics --persona "折腾到吐"

        python -m rag_system analytics --matrix

        python -m rag_system analytics -o output/analytics_report.json
    """
    import json
    from pathlib import Path

    from rag_system.generation.analytics import (
        generate_report, format_report, read_events,
        persona_category_matrix, format_matrix_report,
    )

    if matrix:
        click.echo(f"Building persona x category matrix (last {days} days)...")
        m = persona_category_matrix(days=days)
        click.echo(format_matrix_report(m))
        if output:
            Path(output).write_text(
                json.dumps(m, ensure_ascii=False, indent=2, default=str),
                encoding="utf-8",
            )
            click.echo(f"JSON matrix saved: {output}")
        return

    click.echo(f"Analyzing pipeline activity over the last {days} days...")
    report = generate_report(days=days)

    if persona:
        # Filter events by persona for display
        events = read_events(days=days)
        filtered = [e for e in events if e.get("persona") == persona]
        click.echo(f"Filtered to persona '{persona}': {len(filtered)} events")
        report["total_generations"] = sum(1 for e in filtered if e.get("type") == "generate")
        report["total_storyboards"] = sum(1 for e in filtered if e.get("type") == "storyboard")

    text = format_report(report)
    click.echo(text)

    if output:
        Path(output).write_text(
            json.dumps(report, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        click.echo(f"JSON report saved: {output}")


# ============================================================
# cover — Cover image prompt generator (封面前置)
# ============================================================

@cli.command("cover")
@click.option("--product", "-p", required=True, help="产品名称")
@click.option("--category", "-c", default="",
              help="品类：keyboard / mouse / monitor / laptop / phone / gpu / headphone / desk_chair")
@click.option("--persona", default="折腾到吐", help="人设名称 (默认: 折腾到吐)")
@click.option("--from-brief", default=None, type=click.Path(exists=True),
              help="从Brief文档中提取封面建议")
@click.option("--suggestion", "-s", default="", help="直接给出封面建议文案")
@click.option("--description", "-d", default="", help="补充产品描述")
@click.option("--style", default="douyin_tech_review",
              help="封面风格：douyin_tech_review / bilibili_review / xiaohongshu (默认: douyin_tech_review)")
@click.option("--output", "-o", default=None, type=click.Path(dir_okay=False),
              help="输出路径 (默认: output/covers/{product}-cover-prompt.txt)")
def cover(product, category, persona, from_brief, suggestion, description,
          style, output):
    """Generate a 5-dimension cover image design prompt. 封面前置工作流.

    Expands a cover concept into a full AI image generation prompt
    covering type, palette, rendering, text overlay, and mood.
    Output is ready to paste into Midjourney / DALL-E / Stable Diffusion.

    Examples:

        python -m rag_system cover -p "ROG龙鳞ACE MINI" -c mouse \\
            -s "鼠标悬浮+RGB光圈+黑色背景"

        python -m rag_system cover -p "迈从K20" -c speaker \\
            --from-brief briefs/maicong.txt

        python -m rag_system cover -p "红魔11SPro" -c phone --persona "朋克" \\
            -s "透明探索版背面展示+跑分数字悬浮" -d "电竞手机，RGB风扇，透明背板"
    """
    from pathlib import Path

    from rag_system.generation.cover_generator import (
        generate_cover_prompt, save_cover_prompt,
    )
    from rag_system.utils import sanitize_filename

    # Resolve cover suggestion
    cover_text = suggestion
    if from_brief and not cover_text:
        from rag_system.generation.brief_analyzer import parse_brief
        brief_text = Path(from_brief).read_text(encoding="utf-8")
        analysis = parse_brief(brief_text)
        cover_text = analysis.cover_suggestion
        if cover_text:
            click.echo(f"Extracted cover concept from brief: {cover_text[:80]}...")
        else:
            click.echo("Brief has no cover suggestion. Provide one with --suggestion.", err=True)
            raise SystemExit(1)

    if not cover_text:
        raise click.UsageError("必须提供 --suggestion 或 --from-brief")

    click.echo(f"Generating cover design for: {product}")
    click.echo(f"Style: {style} | Persona: {persona}")

    prompt = generate_cover_prompt(
        product_name=product,
        category=category,
        persona=persona,
        cover_suggestion=cover_text,
        product_description=description,
        style=style,
    )

    out_path = Path(output) if output else None
    saved_path = save_cover_prompt(prompt, product, output_dir=out_path.parent if out_path else None)
    if out_path:
        import shutil
        shutil.move(str(saved_path), str(out_path))
        saved_path = out_path

    click.echo(f"\nCover prompt saved: {saved_path}")
    click.echo("\nTo generate the actual image, paste the prompt into:")
    click.echo("  Midjourney / DALL-E / Stable Diffusion / seedance")
    click.echo(f"\nPreview:\n{prompt[:400]}...")


# ============================================================
# dashboard — Dynamic dashboard generator (数据中台)
# ============================================================

@cli.command("dashboard")
@click.option("--output", "-o", default=None, type=click.Path(dir_okay=False),
              help="输出路径 (默认: output/dashboard.html)")
def dashboard(output):
    """Regenerate dashboard.html with live project statistics. 数据中台.

    Scans knowledge base, output directories, wiki, and git log
    to produce a fully dynamic HTML dashboard with real numbers.

    Example:

        python -m rag_system dashboard
    """
    from pathlib import Path

    from rag_system.generation.dashboard_generator import collect_data, generate_dashboard

    click.echo("Collecting project statistics...")
    data = collect_data()

    click.echo(f"  KB: {data['kb_chunks']} chunks from {data['kb_sources']} sources "
               f"({data['kb_categories']} categories)")
    click.echo(f"  Output: {data['scripts_count']} scripts | "
               f"{data['storyboards_count']} storyboards | "
               f"{data['audits_count']} audits")
    click.echo(f"  Competitive: {data['competitive_count']} videos | "
               f"Wiki: {data['wiki_pages']} pages")
    click.echo(f"  Code: {data['code_lines']} lines | {data['code_modules']} modules | "
               f"{data['git_total_commits']} commits")

    out_path = Path(output) if output else None
    saved = generate_dashboard(data, output_path=out_path)
    click.echo(f"Dashboard saved: {saved}")


# ============================================================
# topic-daily — AI每日选题日报 (柱子哥选题策略)
# ============================================================

@cli.command("topic-daily")
@click.option("--persona", "-p", default="折腾到吐", help="人设名称 (默认: 折腾到吐)")
@click.option("--focus", "-f", default="tech",
              help="关注领域: tech / finance / ai / auto / all (默认: tech)")
@click.option("--top-n", "-n", default=5, type=int, help="返回选题数量 (默认: 5)")
@click.option("--output", "-o", default=None, type=click.Path(dir_okay=False),
              help="保存到文件")
def topic_daily(persona, focus, top_n, output):
    """Generate daily topic brief — AI自动选题日报。

    Scrapes hot topics, scores on 6 dimensions, outputs ranked brief.
    柱子哥方法论: 信息不值钱，观点值钱。

    Example:
        python -m rag_system topic-daily
        python -m rag_system topic-daily --persona "朋克" --focus ai -n 10
    """
    from rag_system.generation.topic_daily import run_topic_daily

    click.echo(f"Generating daily topic brief... (persona={persona}, focus={focus})")
    text = run_topic_daily(persona=persona, category_focus=focus, top_n=top_n, output=output)
    click.echo(text)

    if output:
        click.echo(f"Saved: {output}")


# ============================================================
# outputs — Output management (产出管理)
# ============================================================

@cli.command("outputs")
@click.option("--type", "output_type", default=None, help="script / storyboard / cover / audit")
@click.option("--product", "-p", default=None, help="按产品名过滤")
@click.option("--latest", "-l", is_flag=True, default=False, help="只显示最新一条")
@click.option("--limit", "-n", default=20, type=int, help="显示数量 (默认: 20)")
def outputs(output_type, product, latest, limit):
    """List generated outputs with metadata — 产出管理.

    Reads output/index.jsonl and displays registered artifacts with
    timestamp, type, product name, and file path.

    Examples:

        python -m rag_system outputs

        python -m rag_system outputs --type script -n 10

        python -m rag_system outputs --product "ROG" --type storyboard

        python -m rag_system outputs --latest
    """
    from rag_system.generation.output_manager import list_outputs, get_latest

    if latest:
        entry = get_latest(output_type)
        if entry:
            click.echo(f"[{entry['ts'][:19]}] {entry['type']}: {entry['path']}")
        else:
            click.echo("No outputs found.")
        return

    entries = list_outputs(output_type=output_type, limit=limit, product=product)
    for e in entries:
        click.echo(f"[{e['ts'][:19]}] {e['type']:<10} {e.get('product', ''):<20} {e['path']}")
    click.echo(f"\n{len(entries)} outputs shown.")


if __name__ == "__main__":
    cli()
