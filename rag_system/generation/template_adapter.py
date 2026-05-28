"""Template adapter — reads client reference xlsx/docx and auto-maps columns.

No template library. Client drops a reference file, adapter detects its
column structure and builds a mapping from shot dict fields to output columns.
"""

from pathlib import Path
from openpyxl import load_workbook
from rag_system.utils import logger

# ----- Header text -> shot dict field name fuzzy matching -----
# Each key is a set of known header substrings. First match wins.
HEADER_KEYWORD_MAP: list[tuple[list[str], str]] = [
    (["镜号", "镜头", "镜", "序号", "segment", "shot"], "shot_number"),
    (["时间", "时长", "duration", "time", "秒"], "duration"),
    (["画面描述", "画面", "visual", "image", "图片"], "visual"),
    (["口播", "文案", "voiceover", "vo", "script", "旁白"], "voiceover"),
    (["景别", "运镜", "framing", "camera move", "镜头语言"], "framing"),
    (["花字", "特效", "huazi", "fx", "subtitle"], "huazi"),
    (["音效", "声画", "audio", "sfx", "sound", "声音"], "audio"),
    (["灯光", "机位", "lighting", "camera_setup", "camera setup", "布光"], "lighting_camera"),
    (["备注", "notes", "备注说明", "remark", "说明"], "notes"),
    (["视觉参考", "visual prompt", "concept", "参考图", "AI提示词", "视觉提示", "prompt"], "visual_prompt"),
    (["修改意见", "修改", "revision", "review", "反馈"], None),
]


def detect_header_row(ws, max_scan: int = 10) -> int:
    """Scan the first N rows and return the row number with the most keyword hits.

    Returns 1-based row number. Falls back to row 1 if nothing detected.
    """
    best_row = 1
    best_score = 0

    for r in range(1, min(ws.max_row or max_scan, max_scan) + 1):
        score = 0
        for c in range(1, (ws.max_column or 20) + 1):
            text = str(ws.cell(row=r, column=c).value or "").strip()
            if not text:
                continue
            for keywords, _ in HEADER_KEYWORD_MAP:
                if any(kw in text for kw in keywords):
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_row = r

    logger.info("Header row detected at row %d (score=%d)", best_row, best_score)
    return best_row


def build_column_mapping(headers: list[str]) -> dict[int, str | None]:
    """Map column index -> shot dict field name (or None for non-data columns).

    Args:
        headers: List of cleaned header texts for each column, e.g.
                 ["镜头", "时间/s", "画面描述", "口播&文案", "备注", "修改意见", "修改意见"]

    Returns:
        {0: "shot_number", 1: "duration", 2: "visual", 3: "voiceover",
         4: "notes", 5: None, 6: None}
    """
    mapping: dict[int, str | None] = {}
    for idx, header in enumerate(headers):
        matched = _match_header(header)
        mapping[idx] = matched
        if matched:
            logger.debug("Column %d: '%s' -> %s", idx, header, matched)
        else:
            logger.debug("Column %d: '%s' -> (skip)", idx, header)
    return mapping


def extract_column_config(ws, header_row: int) -> list[dict]:
    """Extract per-column configuration from the reference sheet.

    Returns list of {letter, header, width, font, fill, alignment} per column.
    """
    configs = []
    for c in range(1, (ws.max_column or 20) + 1):
        letter = _col_letter(c)
        cell = ws.cell(row=header_row, column=c)
        header = str(cell.value or "").strip()
        if not header:
            continue
        width = ws.column_dimensions[letter].width or 12
        configs.append({
            "letter": letter,
            "col_index": c,
            "header": header,
            "width": width,
            "font": cell.font,
            "fill": cell.fill,
            "alignment": cell.alignment,
        })
    return configs


def preview_column_mapping(columns: list[str], sample_shot: dict | None = None) -> str:
    """Return a human-readable preview of the column mapping + sample row.

    Used by --preview to verify format before running the full LLM pipeline.
    """
    mapping = build_column_mapping(columns)
    lines = []
    lines.append("")
    lines.append("  列映射预览:")
    lines.append(f"  {'Col':<6} {'列名':<20} {'-> 数据字段':<20}")
    lines.append(f"  {'-'*4:<6} {'-'*18:<20} {'-'*18:<20}")
    for idx, field in mapping.items():
        name = columns[idx] if idx < len(columns) else "?"
        field_display = field or "(留空)"
        lines.append(f"  {idx+1:<6} {name:<20} -> {field_display:<20}")
    lines.append("")

    if sample_shot:
        lines.append("  样例行 (第1镜):")
        for idx, field in mapping.items():
            val = resolve_field_value(sample_shot, field)
            name = columns[idx] if idx < len(columns) else "?"
            display = val[:60] + "..." if len(val) > 60 else val
            lines.append(f"  [{name}] {display}")
        lines.append("")

    return "\n".join(lines)


def columns_to_config(columns: list[str], default_width: float = 18) -> list[dict]:
    """Convert a list of column names into column configs.

    Lightweight path when no reference file is available — the user
    describes columns verbally (e.g., --columns "镜头,时间,画面,口播,备注").
    Each column gets a default width; the mapping is resolved later via
    build_column_mapping().
    """
    configs = []
    for i, name in enumerate(columns):
        name = name.strip()
        if not name:
            continue
        letter = _col_letter(i + 1)
        configs.append({
            "letter": letter,
            "col_index": i + 1,
            "header": name,
            "width": default_width,
            "font": None,
            "fill": None,
            "alignment": None,
        })
    return configs


def resolve_field_value(shot: dict, field: str | None) -> str:
    """Convert a shot dict entry to a cell value based on the mapped field."""
    if field is None:
        return ""
    if field == "shot_number":
        return str(shot.get("shot_number", ""))
    if field == "duration":
        dur = shot.get("duration", "3")
        return str(dur).rstrip("s").strip()
    if field == "visual":
        return str(shot.get("visual", ""))
    if field == "voiceover":
        return str(shot.get("voiceover", ""))
    if field == "framing":
        jingbie = str(shot.get("jingbie", ""))
        yunjing = str(shot.get("yunjing", ""))
        parts = [p for p in [jingbie, yunjing] if p]
        return " | ".join(parts) if parts else ""
    if field == "huazi":
        return str(shot.get("huazi", ""))
    if field == "audio":
        return str(shot.get("audio", ""))
    if field == "lighting_camera":
        lighting = str(shot.get("lighting", ""))
        cam = str(shot.get("camera_setup", ""))
        parts = [p for p in [lighting, cam] if p]
        return "\n".join(parts) if parts else ""
    if field == "notes":
        return str(shot.get("notes", ""))
    if field == "visual_prompt":
        return str(shot.get("visual_prompt", ""))
    return ""


def has_any_mapped_columns(mapping: dict[int, str | None]) -> bool:
    """True if at least one column maps to a data field."""
    return any(v is not None for v in mapping.values())


# ----- internal helpers -----

def _match_header(header: str) -> str | None:
    """Match a single header text to a shot dict field name."""
    for keywords, field_name in HEADER_KEYWORD_MAP:
        if any(kw in header for kw in keywords):
            return field_name
    return None


def _col_letter(index: int) -> str:
    """1 -> 'A', 27 -> 'AA'"""
    result = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result
