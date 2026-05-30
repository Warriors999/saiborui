"""Format storyboard JSON into a professionally styled .xlsx file.

9-column layout optimized for A4 landscape print:
  A:镜号 B:景别·运镜 C:画面描述 D:口播文案 E:时长 F:花字/特效 G:音效/声画 H:灯光/机位 I:备注

Design: Swiss Modernism + Executive Dashboard
  Deep blue #1E40AF header | White + #DBEAFE alternating rows | Slate borders
  Microsoft YaHei 10pt body | 9pt secondary | 12pt title
"""

import math
import re as _re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .template_adapter import (
    detect_header_row,
    build_column_mapping,
    extract_column_config,
    resolve_field_value,
)

# ── Design System Colors ──
PRIMARY = "1E40AF"
PRIMARY_LIGHT = "DBEAFE"
BACKGROUND = "FFFFFF"
SURFACE = "F8FAFC"
BORDER_COLOR = "CBD5E1"
TEXT_PRIMARY = "1E293B"
TEXT_SECONDARY = "64748B"
TEXT_HEADER = "FFFFFF"

# ── Shared Styles (module-level, used by all formatters) ──
thin_side = Side(style="thin", color=BORDER_COLOR)
border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_bottom_thick = Border(left=thin_side, right=thin_side, top=thin_side,
                             bottom=Side(style="medium", color=PRIMARY))
font_body = Font(name="微软雅黑", size=10, color=TEXT_PRIMARY)
font_small = Font(name="微软雅黑", size=9, color=TEXT_SECONDARY)
font_col_header = Font(name="微软雅黑", size=10, bold=True, color=TEXT_HEADER)
font_title_header = Font(name="微软雅黑", size=12, bold=True, color=TEXT_HEADER)
font_shot_number = Font(name="微软雅黑", size=10, bold=True, color=PRIMARY)
fill_header = PatternFill("solid", fgColor=PRIMARY)
fill_alt_row = PatternFill("solid", fgColor=PRIMARY_LIGHT)
fill_white = PatternFill("solid", fgColor=BACKGROUND)
align_center = Alignment(horizontal="center", vertical="center")
align_center_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)
align_left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
align_left_center = Alignment(horizontal="left", vertical="center")


def format_storyboard_to_xlsx(
    storyboard: dict,
    product_name: str,
    persona: str,
    output_path: Path,
    reference_path: Path | None = None,
    columns: list[str] | None = None,
) -> Path:
    """Convert storyboard JSON to a professionally formatted .xlsx file.

    Args:
        reference_path: If provided, delegate to template-based formatting
            using the reference file's column structure.
        columns: If provided (and no reference_path), use these column names
            directly (e.g., ["镜头", "时间", "画面描述", "口播", "备注"]).
    """
    if reference_path:
        return format_storyboard_to_template(
            storyboard, product_name, persona, output_path, reference_path
        )
    if columns:
        return format_storyboard_with_columns(
            storyboard, product_name, persona, output_path, columns
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "分镜脚本"

    metadata = storyboard.get("metadata", {})
    shots = storyboard.get("shots", [])

    # ── Column Widths (9 columns, A4 landscape) ──
    col_widths = {
        "A": 5.0,   # 镜号
        "B": 9.0,   # 景别·运镜
        "C": 38,    # 画面描述
        "D": 48,    # 口播文案
        "E": 5.0,   # 时长
        "F": 17,    # 花字/特效
        "G": 16,    # 音效/声画
        "H": 28,    # 灯光/机位/光比
        "I": 20,    # 备注
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Typography ──
    font_title = Font(name="微软雅黑", size=12, bold=True, color=TEXT_PRIMARY)
    font_title_header = Font(name="微软雅黑", size=12, bold=True, color=TEXT_HEADER)
    font_label_accent = Font(name="微软雅黑", size=10, bold=True, color=PRIMARY)
    font_body = Font(name="微软雅黑", size=10, color=TEXT_PRIMARY)
    font_body_bold = Font(name="微软雅黑", size=10, bold=True, color=TEXT_PRIMARY)
    font_small = Font(name="微软雅黑", size=9, color=TEXT_SECONDARY)
    font_col_header = Font(name="微软雅黑", size=10, bold=True, color=TEXT_HEADER)
    font_shot_number = Font(name="微软雅黑", size=10, bold=True, color=PRIMARY)
    font_framing = Font(name="微软雅黑", size=10, color=TEXT_PRIMARY)

    # ── Fills ──
    fill_header = PatternFill("solid", fgColor=PRIMARY)
    fill_alt_row = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    fill_white = PatternFill("solid", fgColor=BACKGROUND)
    fill_metadata_bg = PatternFill("solid", fgColor=SURFACE)

    # ── Borders ──
    thin_side = Side(style="thin", color=BORDER_COLOR)
    border_all_thin = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side,
    )
    border_bottom_thick = Border(
        left=thin_side, right=thin_side, top=thin_side,
        bottom=Side(style="medium", color=PRIMARY),
    )

    # ── Alignments ──
    align_center = Alignment(horizontal="center", vertical="center")
    align_center_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)
    align_left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_left_center = Alignment(horizontal="left", vertical="center")

    # ═══════════════════════════════════════════
    # METADATA HEADER (Rows 1-7)
    # ═══════════════════════════════════════════
    meta_rows = [
        (1, "分镜脚本", None, font_title, 32, True),
        (2, "达人名称", persona, font_body, 24, False),
        (3, "标题", metadata.get("title", product_name), font_body_bold, 26, False),
        (4, "必带话题", metadata.get("hashtags", ""), font_small, 22, False),
        (5, "内容方向", f"测评+种草          |          视频总时长：{metadata.get('total_duration', '60-90s')}", font_body, 22, False),
        (6, "封面文案", metadata.get("title", product_name), font_small, 22, False),
        (7, "拍摄版式", "16:9 横版视频", font_body, 22, False),
    ]

    for row_num, label, value, val_font, row_height, is_title in meta_rows:
        ws.row_dimensions[row_num].height = row_height
        if is_title:
            ws.merge_cells(f"A{row_num}:I{row_num}")
            cell_label = ws.cell(row=row_num, column=1, value=label)
            cell_label.font = font_title_header
            cell_label.alignment = align_center
            cell_label.fill = fill_header
        else:
            ws.merge_cells(f"A{row_num}:B{row_num}")
            cell_label = ws.cell(row=row_num, column=1, value=f"{label}：")
            cell_label.font = font_label_accent
            cell_label.alignment = Alignment(horizontal="right", vertical="center")
            cell_label.fill = fill_metadata_bg
            ws.merge_cells(f"C{row_num}:I{row_num}")
            cell_val = ws.cell(row=row_num, column=3, value=value)
            cell_val.font = val_font
            cell_val.alignment = align_left_center
            cell_val.fill = fill_metadata_bg

    # Row 8: Separator
    ws.row_dimensions[8].height = 8

    # ═══════════════════════════════════════════
    # COLUMN HEADERS (Row 9)
    # ═══════════════════════════════════════════
    col_headers = [
        "镜号", "景别·运镜", "画面描述", "口播文案", "时长",
        "花字/特效", "音效/声画", "灯光/机位", "备注",
    ]
    ws.row_dimensions[9].height = 28
    for ci, header_text in enumerate(col_headers, 1):
        cell = ws.cell(row=9, column=ci, value=header_text)
        cell.font = font_col_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin

    # ═══════════════════════════════════════════
    # SHOT DATA ROWS (Row 10+)
    # ═══════════════════════════════════════════
    for ri, shot in enumerate(shots):
        row_num = 10 + ri
        is_alt = ri % 2 == 1
        row_fill = fill_alt_row if is_alt else fill_white

        # ── Build each column ──
        shot_num = str(shot.get("shot_number", ri + 1))

        # B: 景别·运镜 (with transition prefix)
        jingbie = shot.get("jingbie", "")
        yunjing = shot.get("yunjing", "")
        transition = shot.get("transition", "")
        framing = " | ".join(p for p in [jingbie, yunjing] if p) or "-"
        if transition and transition not in ("硬切", "开场", ""):
            framing = f"[转场:{transition}] {framing}"

        # C: 画面描述 — pure visual, no camera prefixes
        visual = shot.get("visual", "")

        # D: 口播
        voiceover = shot.get("voiceover", "")

        # E: 时长
        duration = shot.get("duration", "")

        # F: 花字
        huazi = shot.get("huazi", "")

        # G: 音效
        audio = shot.get("audio", "")

        # H: 灯光/机位/光比
        lighting = shot.get("lighting", "")
        camera_setup = shot.get("camera_setup", "")
        light_ratio = _get_light_ratio(shot)
        light_parts = []
        if lighting:
            light_parts.append(lighting)
        if camera_setup:
            light_parts.append(camera_setup)
        # Always include light ratio
        light_parts.append(f"光比: 主:辅≈{light_ratio}")
        light_cam = "\n".join(light_parts)

        # I: 备注 — simple, actionable, ≤20 chars
        notes = shot.get("notes", "")
        notes = _re.sub(r'\[幕:[^\]]+\]', '', notes)
        notes = _re.sub(r'\[转场:[^\]]+\]', '', notes).strip()
        if len(notes) > 20:
            notes = notes[:18] + "…"

        data = [shot_num, framing, visual, voiceover, duration, huazi, audio, light_cam, notes]

        # ── Calculate row height (column-aware) ──
        col_chars_per_line = {
            2: 4.5,   # B: ~4.5 Chinese chars/line at width 9
            3: 19,    # C: ~19 at width 38
            4: 24,    # D: ~24 at width 48
            6: 8.5,   # F: ~8.5 at width 17
            7: 8,     # G: ~8 at width 16
            8: 14,    # H: ~14 at width 28
            9: 10,    # I: ~10 at width 20
        }
        max_lines = 1
        for ci, val in enumerate(data):
            if ci + 1 in col_chars_per_line and val:
                lines_needed = max(1, len(str(val)) / col_chars_per_line[ci + 1])
                max_lines = max(max_lines, lines_needed)
        ws.row_dimensions[row_num].height = max(28, min(140, int(max_lines * 18 + 6)))

        # ── Write cells ──
        for ci, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=ci, value=value)
            cell.fill = row_fill

            if ci == 1:  # 镜号
                cell.font = font_shot_number
                cell.alignment = align_center
            elif ci == 2:  # 景别·运镜
                cell.font = font_framing
                cell.alignment = align_center_wrap
            elif ci == 5:  # 时长
                cell.font = font_body
                cell.alignment = align_center_wrap
            elif ci in (6, 7):  # 花字, 音效
                cell.font = font_small
                cell.alignment = align_left_wrap
            elif ci in (8, 9):  # 灯光/机位, 备注
                cell.font = font_small
                cell.alignment = align_left_wrap
            else:  # 画面描述, 口播
                cell.font = font_body
                cell.alignment = align_left_wrap

            if ri == len(shots) - 1:
                cell.border = border_bottom_thick
            else:
                cell.border = border_all_thin

    # ═══════════════════════════════════════════
    # FINAL TOUCHES — Main Sheet
    # ═══════════════════════════════════════════
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.auto_filter.ref = f"A9:I{9 + len(shots)}"
    ws.freeze_panes = "A10"

    # ═══════════════════════════════════════════
    # LIGHTING DIAGRAM SHEET — per-shot top-down
    # ═══════════════════════════════════════════
    _create_lighting_svgs(wb, shots, product_name, output_path)

    wb.save(str(output_path))
    return output_path


def format_storyboard_to_template(
    storyboard: dict,
    product_name: str,
    persona: str,
    output_path: Path,
    reference_path: Path,
) -> Path:
    """Format storyboard to a template-based xlsx matching the reference file's
    column structure, using the project's Swiss color scheme.

    Workflow:
    1. Load the reference xlsx and detect its header row.
    2. Build a column mapping (header text -> shot dict field).
    3. Extract column widths from the reference.
    4. Write data rows by resolving field values through the mapping.
    5. Apply Swiss-style formatting and freeze panes.
    """
    # ── 1. Load reference & detect structure ──
    ref_wb = load_workbook(reference_path)
    ref_ws = ref_wb.active
    header_row = detect_header_row(ref_ws)

    # Read headers from the detected row
    headers = []
    for c in range(1, (ref_ws.max_column or 20) + 1):
        text = str(ref_ws.cell(row=header_row, column=c).value or "").strip()
        if not text:
            break
        headers.append(text)

    column_mapping = build_column_mapping(headers)
    column_configs = extract_column_config(ref_ws, header_row)

    # Close reference -- we have everything we need
    ref_wb.close()

    # ── 2. Create output workbook ──
    shots = storyboard.get("shots", [])
    wb = Workbook()
    ws = wb.active
    ws.title = "分镜脚本"

    num_cols = len(column_configs)

    # ── 3. Column widths from reference ──
    for config in column_configs:
        ws.column_dimensions[config["letter"]].width = config.get("width", 12)

    # ── 4. Header row ──
    ws.row_dimensions[1].height = 28
    for ci, config in enumerate(column_configs):
        cell = ws.cell(row=1, column=ci + 1, value=config["header"])
        cell.font = font_col_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin

    # ── 5. Data rows ──
    for ri, shot in enumerate(shots):
        row_num = 2 + ri
        is_alt = ri % 2 == 1
        row_fill = fill_alt_row if is_alt else fill_white

        # Write cells via column mapping
        for ci, config in enumerate(column_configs):
            field = column_mapping.get(ci)
            value = resolve_field_value(shot, field)

            cell = ws.cell(row=row_num, column=ci + 1, value=value)
            cell.font = font_body
            cell.fill = row_fill
            cell.alignment = align_left_wrap
            cell.border = border_all_thin

        # Auto-calculate row height
        max_lines = 1
        for ci, config in enumerate(column_configs):
            field = column_mapping.get(ci)
            if field is not None:
                val = resolve_field_value(shot, field)
                if val:
                    col_width = config.get("width", 12)
                    if col_width > 0:
                        char_width = col_width / 2.0  # rough estimate for CJK
                        lines_needed = max(1, len(str(val)) / char_width)
                        max_lines = max(max_lines, lines_needed)
        ws.row_dimensions[row_num].height = max(28, min(140, int(max_lines * 18 + 6)))

    # ── 6. Final touches ──
    ws.freeze_panes = "A2"
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    last_col = get_column_letter(num_cols) if num_cols else "A"
    ws.auto_filter.ref = f"A1:{last_col}{1 + len(shots)}"

    # Lighting diagrams for template mode
    _create_lighting_svgs(wb, shots, product_name, output_path)

    wb.save(str(output_path))
    return output_path


def format_storyboard_with_columns(
    storyboard: dict,
    product_name: str,
    persona: str,
    output_path: Path,
    columns: list[str],
) -> Path:
    """Format storyboard using user-specified column names (no reference file).

    Lightweight variant of format_storyboard_to_template — the user describes
    columns verbally via --columns instead of providing a reference xlsx.
    """
    from .template_adapter import columns_to_config, build_column_mapping, resolve_field_value

    shots = storyboard.get("shots", [])
    configs = columns_to_config(columns)
    headers = [c["header"] for c in configs]
    mapping = build_column_mapping(headers)
    num_cols = len(configs)

    wb = Workbook()
    ws = wb.active
    ws.title = "分镜脚本"

    # Column widths
    for cfg in configs:
        ws.column_dimensions[cfg["letter"]].width = cfg["width"]

    # Header row
    for ci, cfg in enumerate(configs, 1):
        cell = ws.cell(row=1, column=ci, value=cfg["header"])
        cell.font = font_col_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin
    ws.row_dimensions[1].height = 28

    # Data rows
    for ri, shot in enumerate(shots):
        row = ri + 2
        is_alt = ri % 2 == 1
        for ci in range(num_cols):
            field = mapping.get(ci)
            value = resolve_field_value(shot, field)
            cell = ws.cell(row=row, column=ci + 1, value=value)
            cell.font = font_body
            cell.alignment = align_left_wrap if ci >= 2 else align_center_wrap
            cell.border = border_all_thin
            if is_alt:
                cell.fill = fill_alt_row
        # Row height
        max_chars = max(
            len(str(resolve_field_value(shot, mapping.get(ci)))) for ci in range(num_cols)
        )
        ws.row_dimensions[row].height = max(28, min(140, int(max_chars / 40 * 18 + 6)))

    ws.freeze_panes = "A2"
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    last_col = get_column_letter(num_cols) if num_cols else "A"
    ws.auto_filter.ref = f"A1:{last_col}{1 + len(shots)}"

    # Lighting diagrams for columns mode
    _create_lighting_svgs(wb, shots, product_name, output_path)

    wb.save(str(output_path))
    return output_path


def _get_lighting_setup(shot: dict) -> dict:
    """Return minimal lighting setup: key angle, fill (optional), ratio.

    Only 1-2 lights. Variation comes from key angle and ratio, not gear list.
    """
    act = shot.get("act", "")
    jingbie = shot.get("jingbie", "")
    visual = shot.get("visual", "")

    # ── Key light angle varies by shot context ──
    if act == "hook":
        key_angle, ratio, fill = "右前60°", "3:1", "辅光:左侧反光板弱补"
    elif act == "reveal":
        key_angle, ratio, fill = "右前45°", "3:1~4:1", "辅光:左侧补光"
    elif act == "deep_dive" and jingbie in ("特写", "大特写"):
        key_angle, ratio, fill = "正前15°", "2:1", "辅光:底部反光板"
    elif act == "proof":
        key_angle, ratio, fill = "正前10°", "2:1", "辅光:左侧反光板"
    elif act == "cta":
        key_angle, ratio, fill = "右前35°", "2:1~3:1", "辅光:左侧补光"
    elif any(kw in visual for kw in ["俯拍", "顶部", "顶置"]):
        key_angle, ratio, fill = "顶部垂直", "2:1", ""  # single light
    elif any(kw in visual for kw in ["手持", "手部", "POV"]):
        key_angle, ratio, fill = "右前30°", "2:1", ""
    elif any(kw in visual for kw in ["360", "环绕"]):
        key_angle, ratio, fill = "右前45°", "2:1~3:1", "辅光:左后45°"
    elif act == "deep_dive":
        key_angle, ratio, fill = "右前40°", "2:1~3:1", "辅光:左侧补光"
    else:
        key_angle, ratio, fill = "右前40°", "2:1", "辅光:左侧补光"

    return {
        "ratio": ratio,
        "key_angle": key_angle,
        "fill": fill,
        "key": f"主光:{key_angle}+柔光箱 | 色温:5600K",
        "camera": "机位:正面 | 焦段:50mm | 光圈:F2.8",
    }


def _get_light_ratio(shot: dict) -> str:
    return _get_lighting_setup(shot)["ratio"]


def _create_lighting_svgs(wb, shots, product_name, output_path):
    """Add lighting reference sheet to the xlsx (no separate SVG files)."""
    groups = {}
    for si, shot in enumerate(shots):
        lt = shot.get("lighting", "").strip()
        cam = shot.get("camera_setup", "").strip()
        ratio = _get_light_ratio(shot)
        key = (lt, cam, ratio)
        if key not in groups:
            groups[key] = []
        groups[key].append(shot.get("shot_number", si + 1))
    if not groups:
        return
    ws = wb.create_sheet("灯位图")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30
    headers = ["灯位", "适用镜号", "主灯光位", "光比", "机位", "布光技法"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = font_col_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin
    ws.row_dimensions[1].height = 26
    for gi, ((lt, cam, ratio), shot_nums) in enumerate(groups.items()):
        row = gi + 2
        label = f"灯位{chr(65+gi)}"
        shot_ref = ", ".join(str(n) for n in shot_nums[:8])
        ref_shot = shots[shot_nums[0] - 1] if shot_nums else {}
        technique = _derive_technique(ref_shot) if ref_shot else ""
        cam_short = (cam or "50mm F2.8").replace("机位:", "").replace("机位：", "")[:25]
        row_data = [label, shot_ref, lt or "未指定", ratio, cam_short, technique]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = font_body
            cell.alignment = align_left_wrap if ci >= 3 else align_center
            cell.border = border_all_thin
        ws.row_dimensions[row].height = max(28, 16 * (len(lt) // 30 + 1))

def _derive_technique(shot: dict) -> str:
    """Derive professional shooting technique note from shot data. Not fabricated."""
    yunjing = shot.get("yunjing", "")
    jingbie = shot.get("jingbie", "")
    jiandu = shot.get("jiandu", "")
    visual = shot.get("visual", "")

    techniques = []

    # Camera movement technique
    if "推" in yunjing:
        techniques.append("滑轨缓推")
    elif "拉" in yunjing:
        techniques.append("滑轨缓拉")
    elif "摇" in yunjing:
        techniques.append("云台匀速摇摄")
    elif "移" in yunjing:
        techniques.append("滑轨横移跟焦")
    elif "跟" in yunjing:
        techniques.append("手持稳定器跟焦")
    elif "升" in yunjing or "降" in yunjing:
        techniques.append("电动滑轨升降")
    elif "环绕" in yunjing:
        techniques.append("电动转盘或手动环绕")
    elif "微距" in yunjing:
        techniques.append("微距镜头+手动对焦")
    elif "POV" in yunjing:
        techniques.append("头戴或手持POV")
    elif "固定" in yunjing:
        techniques.append("三脚架锁定云台")

    # Angle technique
    if "仰拍" in jiandu:
        techniques.append("低机位仰角")
    elif "俯拍" in jiandu:
        techniques.append("高机位俯角/顶置")
    elif "越肩" in jiandu:
        techniques.append("肩后机位")

    # Shot type technique
    if "特写" in jingbie:
        if "大特写" in jingbie:
            techniques.append("微距镜头")
        else:
            techniques.append("长焦端压缩景深")

    if not techniques:
        techniques.append("标准机位")

    return " | ".join(techniques[:2])
