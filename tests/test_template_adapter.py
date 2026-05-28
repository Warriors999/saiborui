"""Test the column mapping logic from template_adapter (no LLM needed)."""

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from rag_system.generation.template_adapter import (
    build_column_mapping,
    resolve_field_value,
    detect_header_row,
    columns_to_config,
    preview_column_mapping,
    HEADER_KEYWORD_MAP,
)


# ── build_column_mapping ──

def test_build_column_mapping_basic():
    """Maps standard Chinese headers to correct shot dict field names."""
    headers = ["镜头", "时间/s", "画面描述", "口播&文案", "备注"]
    mapping = build_column_mapping(headers)
    expected = {
        0: "shot_number",
        1: "duration",
        2: "visual",
        3: "voiceover",
        4: "notes",
    }
    assert mapping == expected, f"Got: {mapping}"


def test_build_column_mapping_revision_columns():
    """Columns like '修改意见' map to None (skip)."""
    headers = ["修改意见", "修改意见"]
    mapping = build_column_mapping(headers)
    expected = {0: None, 1: None}
    assert mapping == expected, f"Got: {mapping}"


def test_build_column_mapping_mixed():
    """Mixed data and non-data columns."""
    headers = ["镜头", "画面描述", "修改意见", "口播", "花字"]
    mapping = build_column_mapping(headers)
    assert mapping[0] == "shot_number"
    assert mapping[1] == "visual"
    assert mapping[2] is None  # 修改意见
    assert mapping[3] == "voiceover"
    assert mapping[4] == "huazi"


def test_build_column_mapping_empty_list():
    """Empty header list returns empty dict."""
    mapping = build_column_mapping([])
    assert mapping == {}


# ── resolve_field_value ──

def test_resolve_field_value_shot_number():
    shot = {"shot_number": "5"}
    assert resolve_field_value(shot, "shot_number") == "5"


def test_resolve_field_value_duration():
    shot = {"duration": "3s"}
    assert resolve_field_value(shot, "duration") == "3"


def test_resolve_field_value_visual():
    shot = {"visual": "产品主图+外观特写"}
    assert resolve_field_value(shot, "visual") == "产品主图+外观特写"


def test_resolve_field_value_voiceover():
    shot = {"voiceover": "这是一个测试口播。"}
    assert resolve_field_value(shot, "voiceover") == "这是一个测试口播。"


def test_resolve_field_value_framing():
    shot = {"jingbie": "特写", "yunjing": "推"}
    result = resolve_field_value(shot, "framing")
    assert "特写" in result
    assert "推" in result
    assert " | " in result


def test_resolve_field_value_huazi():
    shot = {"huazi": "54g轻量化"}
    assert resolve_field_value(shot, "huazi") == "54g轻量化"


def test_resolve_field_value_audio():
    shot = {"audio": "金属敲击声"}
    assert resolve_field_value(shot, "audio") == "金属敲击声"


def test_resolve_field_value_lighting_camera():
    shot = {"lighting": "主灯:右前45°", "camera_setup": "机位:仰拍30°"}
    result = resolve_field_value(shot, "lighting_camera")
    assert "主灯" in result
    assert "仰拍" in result


def test_resolve_field_value_notes():
    shot = {"notes": "桌面铺黑布"}
    assert resolve_field_value(shot, "notes") == "桌面铺黑布"


def test_resolve_field_value_none_field():
    """None field returns empty string."""
    shot = {"voiceover": "hello"}
    assert resolve_field_value(shot, None) == ""


def test_resolve_field_value_unknown_field():
    """Unknown field name returns empty string."""
    shot = {"voiceover": "hello"}
    assert resolve_field_value(shot, "nonexistent_field") == ""


# ── detect_header_row ──

def test_detect_header_row_basic():
    """detect_header_row finds the row with keywords in a simple mock sheet."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    # Row 1: junk
    ws.cell(row=1, column=1, value="项目信息表")
    ws.cell(row=1, column=2, value="2024Q1")
    # Row 2: actual headers
    ws.cell(row=2, column=1, value="镜头")
    ws.cell(row=2, column=2, value="时间/s")
    ws.cell(row=2, column=3, value="画面描述")
    ws.cell(row=2, column=4, value="口播&文案")
    # Row 3: data
    ws.cell(row=3, column=1, value="1")
    ws.cell(row=3, column=2, value="3")
    ws.cell(row=3, column=3, value="测试画面")

    result = detect_header_row(ws)
    assert result == 2, f"Expected row 2, got {result}"
    wb.close()


def test_detect_header_row_no_keywords():
    """When no headers match, falls back to row 1."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="A")
    ws.cell(row=1, column=2, value="B")
    ws.cell(row=2, column=1, value="X")
    ws.cell(row=2, column=2, value="Y")

    result = detect_header_row(ws)
    assert result == 1  # Falls back to row 1 when nothing matches
    wb.close()


# ── columns_to_config ──

def test_columns_to_config_count():
    """Returns correct number of configs."""
    columns = ["镜头", "时间", "画面", "口播", "备注"]
    configs = columns_to_config(columns)
    assert len(configs) == 5


def test_columns_to_config_default_width():
    """Uses default width when specified."""
    columns = ["镜头", "时间"]
    configs = columns_to_config(columns, default_width=20)
    for cfg in configs:
        assert cfg["width"] == 20


def test_columns_to_config_letter_mapping():
    """First column maps to letter A."""
    columns = ["镜头", "时间", "画面", "口播"]
    configs = columns_to_config(columns)
    assert configs[0]["letter"] == "A"
    assert configs[1]["letter"] == "B"
    assert configs[2]["letter"] == "C"
    assert configs[3]["letter"] == "D"


def test_columns_to_config_empty_strings_skipped():
    """Empty strings in columns are skipped."""
    columns = ["镜头", "", "  ", "口播"]
    configs = columns_to_config(columns)
    # Only non-empty columns after strip
    assert len(configs) == 2


# ── preview_column_mapping ──

def test_preview_column_mapping_non_empty():
    """Returns a non-empty string with column mapping info."""
    columns = ["镜头", "时间/s", "画面描述", "口播&文案", "备注"]
    preview = preview_column_mapping(columns)
    assert preview
    assert isinstance(preview, str)
    assert len(preview) > 0
    assert "列映射预览" in preview


def test_preview_column_mapping_with_sample():
    """With a sample shot, includes sample row output."""
    columns = ["镜头", "时间/s", "画面描述"]
    sample_shot = {
        "shot_number": 1,
        "duration": "3",
        "visual": "产品外观特写",
    }
    preview = preview_column_mapping(columns, sample_shot)
    assert "样例行" in preview
    assert "产品外观特写" in preview


# ── HEADER_KEYWORD_MAP structure ──

def test_header_keyword_map_has_expected_entries():
    """HEADER_KEYWORD_MAP is a list of tuples."""
    assert isinstance(HEADER_KEYWORD_MAP, list)
    assert len(HEADER_KEYWORD_MAP) > 5
    for entry in HEADER_KEYWORD_MAP:
        assert isinstance(entry, tuple)
        assert len(entry) == 2
        assert isinstance(entry[0], list)
        # field_name can be str or None


if __name__ == "__main__":
    test_build_column_mapping_basic()
    test_build_column_mapping_revision_columns()
    test_build_column_mapping_mixed()
    test_build_column_mapping_empty_list()
    test_resolve_field_value_shot_number()
    test_resolve_field_value_duration()
    test_resolve_field_value_visual()
    test_resolve_field_value_voiceover()
    test_resolve_field_value_framing()
    test_resolve_field_value_huazi()
    test_resolve_field_value_audio()
    test_resolve_field_value_lighting_camera()
    test_resolve_field_value_notes()
    test_resolve_field_value_none_field()
    test_resolve_field_value_unknown_field()
    test_detect_header_row_basic()
    test_detect_header_row_no_keywords()
    test_columns_to_config_count()
    test_columns_to_config_default_width()
    test_columns_to_config_letter_mapping()
    test_columns_to_config_empty_strings_skipped()
    test_preview_column_mapping_non_empty()
    test_preview_column_mapping_with_sample()
    test_header_keyword_map_has_expected_entries()
    print("All test_template_adapter tests passed.")
